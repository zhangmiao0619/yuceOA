# -*- coding: utf-8 -*-
"""
油菜小麦面积核对插件
使用方法：将此脚本放到Excel文件同目录，双击运行
"""

import os
import sys
import shutil
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("正在安装依赖，请稍候...")
    os.system('pip install openpyxl')
    import openpyxl

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    from openpyxl.styles import Font, PatternFill, Alignment

def find_excel_file():
    """查找当前目录下的Excel文件"""
    current_dir = Path(__file__).parent
    excel_files = list(current_dir.glob("*.xlsx")) + list(current_dir.glob("*.xls"))
    
    # 排除结果文件
    excel_files = [f for f in excel_files if "核对结果" not in f.name]
    
    if not excel_files:
        return None
    
    if len(excel_files) == 1:
        return excel_files[0]
    
    print("\n找到以下Excel文件，请选择:")
    for i, f in enumerate(excel_files):
        print(f"  {i+1}. {f.name}")
    
    while True:
        try:
            choice = int(input("\n请输入序号: "))
            if 1 <= choice <= len(excel_files):
                return excel_files[choice - 1]
        except ValueError:
            pass

def read_data(file_path):
    """读取Excel数据"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    # 查找字段索引
    field_map = {}
    for i, h in enumerate(headers):
        if h:
            h_lower = str(h).lower()
            if '身份证' in h:
                field_map['id_card'] = i
            elif '姓名' in h:
                field_map['name'] = i
            elif '组别' in h:
                field_map['group'] = i
            elif '地块' in h:
                field_map['parcel'] = i
            elif '油菜' in h and '面积' in h and '最大' not in h:
                field_map['rapeseed_area'] = i
            elif '油菜' in h and '最大' in h:
                field_map['rapeseed_max'] = i
            elif '小麦' in h and '面积' in h and '最大' not in h:
                field_map['wheat_area'] = i
            elif '小麦' in h and '最大' in h:
                field_map['wheat_max'] = i
            elif 'scmjm' in h or ('实测' in h and '面积' in h):
                field_map['scmjm'] = i
    
    required_fields = ['id_card', 'rapeseed_area', 'rapeseed_max', 'wheat_area', 'wheat_max', 'scmjm']
    for field in required_fields:
        if field not in field_map:
            print(f"错误: 缺少必要字段 - {field}")
            return None, None
    
    # 读取数据行
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[field_map['id_card']]:
            data.append({
                'id_card': str(row[field_map['id_card']]).strip(),
                'name': row[field_map.get('name', 0)] or '',
                'group': str(row[field_map.get('group', 0)] or '').strip(),
                'parcel': str(row[field_map.get('parcel', 0)] or '').strip(),
                'rapeseed_area': float(row[field_map['rapeseed_area']] or 0),
                'rapeseed_max': float(row[field_map['rapeseed_max']] or 0),
                'wheat_area': float(row[field_map['wheat_area']] or 0),
                'wheat_max': float(row[field_map['wheat_max']] or 0),
                'scmjm': float(row[field_map['scmjm']] or 0),
            })
    
    return data, field_map

def audit_crop_area(data):
    """核对面积"""
    # 按身份证号分组
    id_card_map = {}
    for row in data:
        id_card = row['id_card']
        if id_card not in id_card_map:
            id_card_map[id_card] = []
        id_card_map[id_card].append(row)
    
    results = []
    insufficient_ids = []
    
    for id_card, rows in id_card_map.items():
        first_row = rows[0]
        has_rapeseed = first_row['rapeseed_area'] > 0 or first_row['rapeseed_max'] > 0
        has_wheat = first_row['wheat_area'] > 0 or first_row['wheat_max'] > 0
        
        total_scmjm = sum(r['scmjm'] for r in rows)
        
        if not has_rapeseed and not has_wheat:
            results.append({
                **first_row,
                'total_scmjm': total_scmjm,
                'audit_type': '',
                'audit_status': '无油菜小麦',
                'required_area': 0,
                'shortage_area': 0,
            })
            continue
        
        if has_rapeseed and has_wheat:
            required = max(first_row['rapeseed_area'], first_row['wheat_area'])
            max_allowed = max(first_row['rapeseed_max'], first_row['wheat_max'])
            audit_type = '油菜' if first_row['rapeseed_max'] >= first_row['wheat_max'] else '小麦'
            
            if total_scmjm > max_allowed:
                results.append({
                    **first_row,
                    'total_scmjm': total_scmjm,
                    'audit_type': audit_type,
                    'audit_status': '面积超限',
                    'required_area': required,
                    'shortage_area': 0,
                })
            elif total_scmjm >= required:
                results.append({
                    **first_row,
                    'total_scmjm': total_scmjm,
                    'audit_type': audit_type,
                    'audit_status': '面积够',
                    'required_area': required,
                    'shortage_area': 0,
                })
            else:
                shortage = required - total_scmjm
                results.append({
                    **first_row,
                    'total_scmjm': total_scmjm,
                    'audit_type': audit_type,
                    'audit_status': '面积不够',
                    'required_area': required,
                    'shortage_area': shortage,
                })
                insufficient_ids.append(id_card)
        elif has_rapeseed:
            required = first_row['rapeseed_area']
            max_allowed = first_row['rapeseed_max']
            
            if total_scmjm > max_allowed:
                results.append({
                    **first_row,
                    'total_scmjm': total_scmjm,
                    'audit_type': '油菜',
                    'audit_status': '面积超限',
                    'required_area': required,
                    'shortage_area': 0,
                })
            elif total_scmjm >= required:
                results.append({
                    **first_row,
                    'total_scmjm': total_scmjm,
                    'audit_type': '油菜',
                    'audit_status': '面积够',
                    'required_area': required,
                    'shortage_area': 0,
                })
            else:
                shortage = required - total_scmjm
                results.append({
                    **first_row,
                    'total_scmjm': total_scmjm,
                    'audit_type': '油菜',
                    'audit_status': '面积不够',
                    'required_area': required,
                    'shortage_area': shortage,
                })
                insufficient_ids.append(id_card)
        elif has_wheat:
            required = first_row['wheat_area']
            max_allowed = first_row['wheat_max']
            
            if total_scmjm > max_allowed:
                results.append({
                    **first_row,
                    'total_scmjm': total_scmjm,
                    'audit_type': '小麦',
                    'audit_status': '面积超限',
                    'required_area': required,
                    'shortage_area': 0,
                })
            elif total_scmjm >= required:
                results.append({
                    **first_row,
                    'total_scmjm': total_scmjm,
                    'audit_type': '小麦',
                    'audit_status': '面积够',
                    'required_area': required,
                    'shortage_area': 0,
                })
            else:
                shortage = required - total_scmjm
                results.append({
                    **first_row,
                    'total_scmjm': total_scmjm,
                    'audit_type': '小麦',
                    'audit_status': '面积不够',
                    'required_area': required,
                    'shortage_area': shortage,
                })
                insufficient_ids.append(id_card)
    
    return results, insufficient_ids

def find_supplements(data, insufficient_ids):
    """查找补充方案"""
    if not insufficient_ids:
        return []
    
    # 分类：需要补充的 和 可以补充的
    id_card_map = {}
    for row in data:
        id_card = row['id_card']
        if id_card not in id_card_map:
            id_card_map[id_card] = row
    
    # 找出无油菜小麦的农户
    no_crop = {}  # group -> list
    need_supplement = {}  # id_card -> {shortage, type, max}
    
    for id_card in id_card_map:
        row = id_card_map[id_card]
        has_rapeseed = row['rapeseed_area'] > 0 or row['rapeseed_max'] > 0
        has_wheat = row['wheat_area'] > 0 or row['wheat_max'] > 0
        
        if id_card in insufficient_ids:
            total_scmjm = sum(r['scmjm'] for r in data if r['id_card'] == id_card)
            required = max(row['rapeseed_area'], row['wheat_area'])
            max_allowed = max(row['rapeseed_max'], row['wheat_max'])
            shortage = required - total_scmjm
            
            if shortage > 0:
                supplement_type = '油菜' if (has_rapeseed and (not has_wheat or row['rapeseed_max'] >= row['wheat_max'])) else '小麦'
                need_supplement[id_card] = {
                    'shortage': shortage,
                    'type': supplement_type,
                    'max': max_allowed,
                    'group': row['group']
                }
        elif not has_rapeseed and not has_wheat:
            group = row['group'] or '未知'
            if group not in no_crop:
                no_crop[group] = []
            no_crop[group].append(row)
    
    supplements = []
    
    for id_card, need in need_supplement.items():
        target_group = need['group']
        
        # 优先从本组补充
        available = no_crop.get(target_group, [])
        remaining = need['shortage']
        
        for row in available:
            if remaining <= 0:
                break
            max_can_add = need['max'] - row['scmjm']
            if max_can_add <= 0:
                continue
            
            add_amount = min(remaining, max_can_add)
            supplements.append({
                'id_card': row['id_card'],
                'name': row['name'],
                'group': row['group'],
                'parcel': row['parcel'],
                'supplement_type': need['type'],
                'supplement_area': add_amount,
                'total_area': row['scmjm'] + add_amount,
                'max_allowed': need['max'],
                'target_id_card': id_card,
            })
            remaining -= add_amount
        
        # 本组不够，从其他组补充
        if remaining > 0:
            for group, rows in no_crop.items():
                if group == target_group:
                    continue
                if remaining <= 0:
                    break
                for row in rows:
                    if remaining <= 0:
                        break
                    max_can_add = need['max'] - row['scmjm']
                    if max_can_add <= 0:
                        continue
                    
                    add_amount = min(remaining, max_can_add)
                    supplements.append({
                        'id_card': row['id_card'],
                        'name': row['name'],
                        'group': row['group'],
                        'parcel': row['parcel'],
                        'supplement_type': need['type'],
                        'supplement_area': add_amount,
                        'total_area': row['scmjm'] + add_amount,
                        'max_allowed': need['max'],
                        'target_id_card': id_card,
                    })
                    remaining -= add_amount
    
    return supplements

def write_results(data, results, supplements, output_path):
    """写入结果Excel"""
    wb = openpyxl.Workbook()
    
    # 核对结果sheet
    ws1 = wb.active
    ws1.title = "核对结果"
    
    headers1 = ['身份证号', '姓名', '组别', '地块', '油菜面积', '油菜最大值', '小麦面积', '小麦最大值', '实测面积总和', '核对类型', '核对状态', '要求面积', '缺少面积']
    ws1.append(headers1)
    
    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    for row in results:
        ws1.append([
            row['id_card'],
            row['name'],
            row['group'],
            row['parcel'],
            row['rapeseed_area'],
            row['rapeseed_max'],
            row['wheat_area'],
            row['wheat_max'],
            row['total_scmjm'],
            row['audit_type'],
            row['audit_status'],
            row['required_area'],
            row['shortage_area'],
        ])
        
        # 设置状态颜色
        last_row = ws1.max_row
        status_cell = ws1.cell(last_row, 11)  # 核对状态列
        
        if row['audit_status'] == '面积够':
            status_cell.fill = good_fill
        elif row['audit_status'] == '面积不够':
            status_cell.fill = bad_fill
    
    # 自动调整列宽
    for col in ws1.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws1.column_dimensions[col_letter].width = min(max_length + 2, 20)
    
    # 补充建议sheet
    if supplements:
        ws2 = wb.create_sheet("补充建议")
        
        headers2 = ['被补充农户', '补充农户', '姓名', '组别', '地块', '补充种类', '补充面积', '补充后总面积', '允许最大值']
        ws2.append(headers2)
        
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for s in supplements:
            ws2.append([
                s['target_id_card'],
                s['id_card'],
                s['name'],
                s['group'],
                s['parcel'],
                s['supplement_type'],
                s['supplement_area'],
                s['total_area'],
                s['max_allowed'],
            ])
        
        for col in ws2.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws2.column_dimensions[col_letter].width = min(max_length + 2, 20)
    
    wb.save(output_path)

def main():
    print("="*50)
    print("     油菜小麦面积核对插件")
    print("="*50)
    
    # 查找Excel文件
    excel_file = find_excel_file()
    
    if not excel_file:
        print("\n未找到Excel文件！")
        print("请将数据文件和此脚本放在同一文件夹")
        input("\n按回车键退出...")
        return
    
    print(f"\n读取文件: {excel_file.name}")
    
    # 读取数据
    data, field_map = read_data(excel_file)
    
    if not data:
        print("数据读取失败！")
        input("\n按回车键退出...")
        return
    
    print(f"共读取 {len(data)} 条数据")
    
    # 核对
    print("\n正在核对...")
    results, insufficient_ids = audit_crop_area(data)
    
    # 统计
    sufficient = sum(1 for r in results if r['audit_status'] == '面积够')
    insufficient = sum(1 for r in results if r['audit_status'] == '面积不够')
    no_crop = sum(1 for r in results if r['audit_status'] == '无油菜小麦')
    
    print(f"\n核对完成!")
    print(f"  - 面积够: {sufficient} 户")
    print(f"  - 面积不够: {insufficient} 户")
    print(f"  - 无油菜小麦: {no_crop} 户")
    
    # 补充建议
    supplements = []
    if insufficient_ids:
        print(f"\n正在计算补充方案...")
        supplements = find_supplements(data, insufficient_ids)
        print(f"  找到 {len(supplements)} 条补充建议")
    
    # 保存结果
    output_file = excel_file.stem + "_核对结果.xlsx"
    output_path = excel_file.parent / output_file
    
    write_results(data, results, supplements, output_path)
    
    print(f"\n结果已保存到: {output_file}")
    print("\n" + "="*50)
    
    input("\n按回车键退出...")

if __name__ == "__main__":
    main()
